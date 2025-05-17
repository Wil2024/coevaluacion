import streamlit as st
import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración del archivo Excel
ARCHIVO_EXCEL = "coevaluaciones.xlsx"

# Crear archivo Excel si no existe
if not os.path.exists(ARCHIVO_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluaciones"
    ws.append(["Equipo", "Estudiante", "Evaluador", "Nota"])
    wb.save(ARCHIVO_EXCEL)

# Función para guardar evaluaciones
def guardar_evaluacion(datos):
    df_nuevos = pd.DataFrame(datos)
    book = load_workbook(ARCHIVO_EXCEL)
    writer = pd.ExcelWriter(ARCHIVO_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    df_existente = pd.read_excel(ARCHIVO_EXCEL, sheet_name="Evaluaciones")
    df_final = pd.concat([df_existente, df_nuevos], ignore_index=True)
    df_final.to_excel(writer, sheet_name="Evaluaciones", index=False)
    writer.close()

# Datos predefinidos de equipos y estudiantes
equipos_estudiantes = {
    "Equipo 1": ["Raul Olaechea", "Fressia", "Paola Errea"],
    "Equipo 2": ["Fiorella Valdivia", "Karla Elescano", "Patricia Sinclair", "Mauricio Negrón"],
    "Equipo 3": ["Alessandra Lavado", "Ericsson Castro", "Antonio Monzón", "Elisabeth Chamorro"],
    "Equipo 4": ["Nina Llamoca", "Elcy Maguiña", "Melany Zevallos", "Javier García","José Tipacti"]
}

# Contraseña del docente
CLAVE_DOCENTE = "docentejwts123"

# Interfaz principal
st.title("🎓 Aplicación de Coevaluación Grupal")

modo = st.sidebar.selectbox("Seleccione modo", ["Estudiante", "Docente"])

if modo == "Estudiante":
    st.header("📝 Formulario de Coevaluación")

    equipo_seleccionado = st.selectbox("Selecciona tu equipo", options=list(equipos_estudiantes.keys()))

    if equipo_seleccionado:
        integrantes = equipos_estudiantes[equipo_seleccionado]
        evaluador = st.selectbox("Tu Nombre", options=integrantes)

        st.write("### Califica a cada compañero (incluyéndote):")
        notas = {}
        for nombre in integrantes:
            nota = st.slider(f"Nota para {nombre}", min_value=0.0, max_value=20.0, step=0.5, key=f"nota_{nombre}")
            notas[nombre] = nota

        if st.button("Enviar Evaluación"):
            if not notas.get(evaluador):
                st.error("Debes calificarte a ti mismo.")
            else:
                datos = []
                for estudiante, nota in notas.items():
                    datos.append({
                        "Equipo": equipo_seleccionado,
                        "Estudiante": estudiante,
                        "Evaluador": evaluador,
                        "Nota": nota
                    })
                guardar_evaluacion(datos)
                st.success("✅ Evaluación enviada correctamente.")

elif modo == "Docente":
    st.header("🔐 Acceso al Modo Docente")
    clave_ingresada = st.text_input("Ingrese la contraseña del docente", type="password")
    if st.button("Ingresar"):
        if clave_ingresada == CLAVE_DOCENTE:
            st.session_state["acceso_docente"] = True
        else:
            st.error("Contraseña incorrecta.")

    if st.session_state.get("acceso_docente", False):
    	st.success("🔓 Acceso concedido al modo docente.")

    	try:
        	df = pd.read_excel(ARCHIVO_EXCEL, sheet_name="Evaluaciones")
    	except Exception as e:
        	st.error("Error al leer el archivo Excel.")
        	df = pd.DataFrame()

    	if df.empty:
        	st.info("No hay evaluaciones registradas aún.")
    	else:
        	st.subheader("Todas las Evaluaciones")
        	st.dataframe(df)

        	st.subheader("Promedio por Estudiante")

        	# Calcular promedio por estudiante (escala 0-20)
        	promedios = df.groupby("Estudiante")["Nota"].mean().round(2).reset_index()
        	promedios.rename(columns={"Nota": "Nota Promedio"}, inplace=True)

        	# Calcular factor de ajuste (promedio / 20)
        	promedios["Factor Ajuste"] = (promedios["Nota Promedio"] / 20).round(2)

        	# Mostrar tabla con ambos valores
        	st.dataframe(promedios)

        	st.download_button(
            		label="📥 Descargar datos",
            		data=open(ARCHIVO_EXCEL, "rb").read(),
            		file_name="coevaluaciones.xlsx",
            		mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        	)

        	if st.button("Cerrar Sesión"):
            		st.session_state["acceso_docente"] = False
            		st.experimental_rerun()
    else:
        st.info("Esperando contraseña...")